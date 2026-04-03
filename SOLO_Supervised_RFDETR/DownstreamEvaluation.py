#!/usr/bin/env python
"""
Spreadsheet-driven downstream quality evaluation on full-FOV images.

Workflow:
- read a manual label file (.xlsx or .csv) with image name/path in column 1 and label in column 2
- run SAHI-based RF-DETR inference on the listed images
- count Leucocyte and Squamous Epithelial Cell detections
- convert counts to downstream quality class (1/2/3)
- compare manual vs predicted downstream class and export metrics/overlays
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import re
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET

try:
    import numpy as np
except Exception:
    np = None  # type: ignore[assignment]

try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:
    Image = None  # type: ignore[assignment]
    ImageDraw = None  # type: ignore[assignment]
    ImageFont = None  # type: ignore[assignment]

try:
    from sklearn.metrics import accuracy_score, balanced_accuracy_score, confusion_matrix, f1_score, precision_recall_fscore_support
except Exception:
    accuracy_score = None  # type: ignore[assignment]
    balanced_accuracy_score = None  # type: ignore[assignment]
    confusion_matrix = None  # type: ignore[assignment]
    f1_score = None  # type: ignore[assignment]
    precision_recall_fscore_support = None  # type: ignore[assignment]


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import EvalRFDETR_SOLO_LocalOnly as local_eval
import EvalRFDETR_SAHI_FOV_Local as fov_eval


ACTIVE_PRESET = "two_class"
DEFAULT_OUTPUT_DIR = r"C:\Users\SH37YE\Desktop\PhD_Code_github\AIPoweredMicroscope\DownstreamOutput"
DEFAULT_IMAGES_ROOT = fov_eval.DEFAULT_IMAGES_ROOT
DEFAULT_MANUAL_LABELS = (
    r"D:\PHD\Results\Quality Assessment\Epi+Leu for ESCMID Conference\FullFOVEval\FullFOVClassification.xlsx"
)
DEFAULT_SHEET_NAME = ""
OVERLAY_SELECTION = "all"
CROSS_CLASS_DUPLICATE_SUPPRESSION = True
CROSS_CLASS_DUPLICATE_IOS_THRESHOLD = 0.75
CROSS_CLASS_DUPLICATE_IOU_THRESHOLD = 0.35
CROSS_CLASS_DUPLICATE_AREA_RATIO_THRESHOLD = 0.55

DOWNSTREAM_LABELS = {
    1: "Qualified",
    2: "Partially qualified",
    3: "Not qualified",
}

DOWNSTREAM_PRESETS = {
    "two_class": {
        "target_name": "Downstream",
        "checkpoint": (
            r"D:\PHD\Results\Quality Assessment\Epi+Leu for ESCMID Conference\third full"
            r"\session_20260327_095556\TwoClass\HPO_Config_001\checkpoint_best_total.pth"
        ),
        "images_root": DEFAULT_IMAGES_ROOT,
        "model_class": "auto",
        "class_names": fov_eval.DEFAULT_CLASS_NAMES,
        "class_score_thresholds": fov_eval.CLASS_SCORE_THRESHOLDS,
    },
}


@dataclass
class DownstreamConfig:
    preset_name: str
    checkpoint: Path
    images_root: Path
    manual_labels_path: Path
    sheet_name: Optional[str]
    output_dir: Path
    output_root_dir: Path
    run_timestamp: str
    target_name: str
    class_names: List[str]
    class_score_thresholds: Dict[str, float]
    model_class: str
    model_resolution: Optional[int]
    score_floor: float
    score_threshold: float
    overlay_selection: str
    num_overlays: int
    seed: int
    progress_every: int
    slice_height: Optional[int]
    slice_width: Optional[int]
    overlap_height_ratio: float
    overlap_width_ratio: float
    perform_standard_pred: bool
    postprocess_type: str
    postprocess_match_metric: str
    postprocess_match_threshold: float
    postprocess_class_agnostic: bool
    cross_class_duplicate_suppression: bool
    cross_class_duplicate_ios_threshold: float
    cross_class_duplicate_iou_threshold: float
    cross_class_duplicate_area_ratio_threshold: float


def _optional_path(raw: str) -> Optional[Path]:
    raw = (raw or "").strip()
    if not raw:
        return None
    return Path(raw).expanduser()


def _resolve_preset(name: str) -> Dict[str, Any]:
    key = (name or "").strip().lower()
    if key not in DOWNSTREAM_PRESETS:
        raise ValueError(
            f"Unknown preset {name!r}. Available presets: {', '.join(sorted(DOWNSTREAM_PRESETS))}"
        )
    return DOWNSTREAM_PRESETS[key]


def ensure_deps() -> None:
    local_eval.ensure_deps()
    if np is None:
        raise ImportError("numpy is required.")
    if Image is None or ImageDraw is None or ImageFont is None:
        raise ImportError("Pillow is required.")


def json_dump(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def build_output_dir(base_output_dir: Path, target_name: str) -> Tuple[Path, Path, str]:
    suffix = re.sub(r"[^A-Za-z0-9]+", "", target_name) or "Downstream"
    root_dir = base_output_dir.parent / f"{base_output_dir.name}_{suffix}"
    run_stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = root_dir / run_stamp
    counter = 2
    while out_dir.exists():
        out_dir = root_dir / f"{run_stamp}_{counter:02d}"
        counter += 1
    return out_dir, root_dir, out_dir.name


def parse_manual_label(value: Any) -> int:
    if value is None:
        raise ValueError("Manual label is empty.")
    txt = str(value).strip()
    if txt.isdigit():
        label_id = int(txt)
        if label_id in DOWNSTREAM_LABELS:
            return label_id

    normalized = re.sub(r"[^a-z0-9]+", "", txt.lower())
    aliases = {
        "1": 1,
        "qualified": 1,
        "velegnet": 1,
        "2": 2,
        "partiallyqualified": 2,
        "partlyqualified": 2,
        "delvistegnet": 2,
        "delvisegnet": 2,
        "3": 3,
        "notqualified": 3,
        "uegnet": 3,
    }
    if normalized in aliases:
        return aliases[normalized]
    raise ValueError(f"Unsupported manual label value: {value!r}")


def is_manual_label_header(first_value: Any, second_value: Any) -> bool:
    first_txt = str(first_value or "").strip().lower()
    second_txt = str(second_value or "").strip().lower()
    first_headers = {"filename", "file_name", "image", "image_name", "path", "image_path"}
    second_headers = {"manual_class", "label", "manual_label", "classification", "class"}
    return first_txt in first_headers or second_txt in second_headers


def _xlsx_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in str(cell_ref) if ch.isalpha()).upper()
    value = 0
    for ch in letters:
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return max(0, value - 1)


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []
    root = ET.fromstring(zf.read(path))
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    values: List[str] = []
    for si in root.findall("a:si", ns):
        texts = [t.text or "" for t in si.findall(".//a:t", ns)]
        values.append("".join(texts))
    return values


def _xlsx_sheet_target(zf: zipfile.ZipFile, sheet_name: Optional[str]) -> str:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    ns_wb = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    rel_map: Dict[str, str] = {}
    for rel in rels:
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rel_map[rel_id] = target

    selected_rel_id: Optional[str] = None
    for sheet in workbook.findall("a:sheets/a:sheet", ns_wb):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if selected_rel_id is None:
            selected_rel_id = rel_id
        if sheet_name and name == sheet_name:
            selected_rel_id = rel_id
            break

    if not selected_rel_id or selected_rel_id not in rel_map:
        raise ValueError(f"Could not resolve worksheet {sheet_name!r} in {zf.filename}")
    target = rel_map[selected_rel_id].replace("\\", "/")
    if not target.startswith("xl/"):
        target = f"xl/{target}"
    return target


def _xlsx_cell_value(cell: ET.Element, shared_strings: Sequence[str]) -> str:
    value_node = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    inline_is = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")
    cell_type = cell.attrib.get("t", "")

    if inline_is is not None:
        texts = [t.text or "" for t in inline_is.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
        return "".join(texts)

    if value_node is None:
        return ""
    raw = value_node.text or ""
    if cell_type == "s":
        idx = int(raw) if raw.strip() else -1
        return shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
    return raw


def load_manual_rows_from_xlsx_fallback(path: Path, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with zipfile.ZipFile(path, "r") as zf:
        shared_strings = _xlsx_shared_strings(zf)
        sheet_target = _xlsx_sheet_target(zf, sheet_name)
        root = ET.fromstring(zf.read(sheet_target))

    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for row_idx, row in enumerate(root.findall(".//a:sheetData/a:row", ns), start=1):
        values: Dict[int, str] = {}
        for cell in row.findall("a:c", ns):
            col_idx = _xlsx_col_index(cell.attrib.get("r", ""))
            values[col_idx] = _xlsx_cell_value(cell, shared_strings)

        first = values.get(0, "")
        second = values.get(1, "")
        if row_idx == 1:
            if is_manual_label_header(first, second):
                continue
        if str(first or "").strip() == "":
            continue
        rows.append(
            {
                "source_value": str(first).strip(),
                "manual_label_id": parse_manual_label(second),
            }
        )
    return rows


def load_manual_rows(path: Path, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                if not row:
                    continue
                if row_idx == 1 and row and is_manual_label_header(row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None):
                    continue
                if len(row) < 2 or not str(row[0]).strip():
                    continue
                rows.append(
                    {
                        "source_value": str(row[0]).strip(),
                        "manual_label_id": parse_manual_label(row[1]),
                    }
                )
        return rows

    if suffix != ".xlsx":
        raise ValueError("Manual label file must be .xlsx or .csv")

    try:
        from openpyxl import load_workbook
    except Exception:
        return load_manual_rows_from_xlsx_fallback(path, sheet_name)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first = row[0] if len(row) >= 1 else None
        second = row[1] if len(row) >= 2 else None
        if row_idx == 1:
            if is_manual_label_header(first, second):
                continue
        if first is None or str(first).strip() == "":
            continue
        rows.append(
            {
                "source_value": str(first).strip(),
                "manual_label_id": parse_manual_label(second),
            }
        )
    wb.close()
    return rows


def build_image_index(images_root: Path) -> Dict[str, List[Path]]:
    sample_dirs = fov_eval.discover_sample_dirs(images_root, 0, None)
    image_paths = fov_eval.discover_images(sample_dirs)
    index: Dict[str, List[Path]] = {}
    for path in image_paths:
        rel = fov_eval.safe_relpath(path, images_root)
        for key in {path.name.lower(), rel.lower()}:
            index.setdefault(key, []).append(path)
    return index


def resolve_image_path(source_value: str, images_root: Path, image_index: Dict[str, List[Path]]) -> Path:
    raw = source_value.strip()
    raw_path = Path(raw)
    if raw_path.is_absolute() and raw_path.exists():
        return raw_path

    candidate = images_root / raw
    if candidate.exists():
        return candidate.resolve()

    key = raw.replace("\\", "/").lower()
    if key in image_index and len(image_index[key]) == 1:
        return image_index[key][0]

    name_key = Path(raw).name.lower()
    if name_key in image_index and len(image_index[name_key]) == 1:
        return image_index[name_key][0]
    if name_key in image_index and len(image_index[name_key]) > 1:
        raise FileNotFoundError(
            f"Image name {raw!r} is ambiguous under {images_root}. Use a relative path instead of just the filename."
        )

    raise FileNotFoundError(f"Could not resolve image path for {raw!r} under {images_root}")


def count_class_predictions(pred_cls: np.ndarray, class_names: Sequence[str]) -> Dict[str, int]:
    counts = {str(name): 0 for name in class_names}
    for cls_idx in pred_cls.tolist():
        if 0 <= int(cls_idx) < len(class_names):
            counts[class_names[int(cls_idx)]] += 1
    return counts


def box_area_xyxy(box: np.ndarray) -> float:
    x1, y1, x2, y2 = [float(v) for v in box.tolist()]
    return max(0.0, x2 - x1) * max(0.0, y2 - y1)


def box_ios_xyxy(box1: np.ndarray, box2: np.ndarray) -> float:
    x11, y11, x12, y12 = [float(v) for v in box1.tolist()]
    x21, y21, x22, y22 = [float(v) for v in box2.tolist()]
    inter_x1 = max(x11, x21)
    inter_y1 = max(y11, y21)
    inter_x2 = min(x12, x22)
    inter_y2 = min(y12, y22)
    inter_w = max(0.0, inter_x2 - inter_x1)
    inter_h = max(0.0, inter_y2 - inter_y1)
    inter = inter_w * inter_h
    smaller = min(box_area_xyxy(box1), box_area_xyxy(box2))
    return inter / smaller if smaller > 0 else 0.0


def box_iou_xyxy(box1: np.ndarray, box2: np.ndarray) -> float:
    iou = local_eval.iou_matrix(box1.reshape(1, 4), box2.reshape(1, 4))
    return float(iou[0, 0]) if iou.size else 0.0


def box_area_ratio(box1: np.ndarray, box2: np.ndarray) -> float:
    area1 = box_area_xyxy(box1)
    area2 = box_area_xyxy(box2)
    larger = max(area1, area2)
    smaller = min(area1, area2)
    return smaller / larger if larger > 0 else 0.0


def suppress_cross_class_duplicates(
    pred_boxes: np.ndarray,
    pred_scores: np.ndarray,
    pred_cls: np.ndarray,
    class_names: Sequence[str],
    ios_threshold: float,
    iou_threshold: float,
    area_ratio_threshold: float,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, int]:
    if len(pred_boxes) <= 1:
        return pred_boxes, pred_scores, pred_cls, 0

    keep = np.ones((len(pred_boxes),), dtype=bool)
    order = np.argsort(-pred_scores)
    suppressed = 0

    for pos_i, i in enumerate(order):
        if not keep[i]:
            continue
        for j in order[pos_i + 1:]:
            if not keep[j]:
                continue
            if int(pred_cls[i]) == int(pred_cls[j]):
                continue

            area_ratio = box_area_ratio(pred_boxes[i], pred_boxes[j])
            if area_ratio < area_ratio_threshold:
                continue

            ios = box_ios_xyxy(pred_boxes[i], pred_boxes[j])
            if ios < ios_threshold:
                continue

            iou = box_iou_xyxy(pred_boxes[i], pred_boxes[j])
            if iou < iou_threshold:
                continue

            keep[j] = False
            suppressed += 1

    return pred_boxes[keep], pred_scores[keep], pred_cls[keep], suppressed


def leucocyte_score(leu_count: int) -> int:
    if leu_count < 10:
        return -1
    if leu_count <= 25:
        return 0
    if leu_count <= 50:
        return 1
    return 2


def squamous_epithelial_score(epi_count: int) -> int:
    if epi_count < 10:
        return 0
    if epi_count <= 25:
        return -1
    return -2


def classify_quality_from_counts(leu_count: int, epi_count: int) -> Tuple[int, int, int, int]:
    leu_score = leucocyte_score(leu_count)
    epi_score = squamous_epithelial_score(epi_count)
    total_score = leu_score + epi_score

    if total_score >= 1 and epi_count < 10:
        label_id = 1
    elif total_score <= -1:
        label_id = 3
    else:
        label_id = 2
    return label_id, leu_score, epi_score, total_score


def draw_downstream_overlay(
    image_path: Path,
    pred_boxes: np.ndarray,
    pred_scores: np.ndarray,
    pred_cls: np.ndarray,
    class_names: Sequence[str],
    leu_count: int,
    epi_count: int,
    manual_label_name: str,
    predicted_label_name: str,
    output_path: Path,
) -> None:
    img = Image.open(image_path).convert("RGB")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    colors = fov_eval.class_color_map(class_names)

    for box, score, cls_idx in zip(pred_boxes, pred_scores, pred_cls):
        color = colors.get(int(cls_idx), (255, 0, 0))
        x1, y1, x2, y2 = [int(round(v)) for v in box.tolist()]
        draw.rectangle([x1, y1, x2, y2], outline=color, width=2)
        text = fov_eval.score_to_text(float(score))
        text_w, text_h = fov_eval._measure_text(draw, text, font)
        tx, ty = fov_eval._score_anchor(box.tolist(), img.width, img.height, text_w, text_h)
        fov_eval.draw_text_with_outline(draw, (tx, ty), text, fill=color, font=font)

    info_lines = [
        f"Leucocyte count: {leu_count}",
        f"Sq. epithelial count: {epi_count}",
        f"Manual: {manual_label_name}",
        f"Predicted: {predicted_label_name}",
    ]
    legend_lines = [f"{class_names[i]} = prediction" for i in range(len(class_names))]
    all_lines = info_lines + legend_lines
    pad = 6
    line_h = max(fov_eval._measure_text(draw, "Ag", font)[1], 10) + 2
    panel_w = max((fov_eval._measure_text(draw, line, font)[0] for line in all_lines), default=0) + 30
    panel_h = len(all_lines) * line_h + pad * 2
    draw.rectangle([6, 6, 6 + panel_w + pad * 2, 6 + panel_h], fill=(0, 0, 0))

    for idx, line in enumerate(info_lines):
        y = 6 + pad + idx * line_h
        fov_eval.draw_text_with_outline(draw, (12, y), line, fill=(255, 255, 255), font=font)
    for idx, line in enumerate(legend_lines, start=len(info_lines)):
        y = 6 + pad + idx * line_h
        color = colors.get(idx - len(info_lines), (255, 0, 0))
        draw.rectangle([12, y + 2, 24, y + line_h - 2], outline=color, width=2)
        fov_eval.draw_text_with_outline(draw, (30, y), line, fill=(255, 255, 255), font=font)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(output_path)


def compute_downstream_metrics(manual_ids: Sequence[int], pred_ids: Sequence[int]) -> Tuple[Dict[str, Any], List[Dict[str, Any]], np.ndarray]:
    ordered_ids = [1, 2, 3]
    labels = [DOWNSTREAM_LABELS[i] for i in ordered_ids]
    if confusion_matrix is None or accuracy_score is None or balanced_accuracy_score is None or precision_recall_fscore_support is None or f1_score is None:
        raise ImportError("scikit-learn is required for downstream metrics.")

    cm = confusion_matrix(manual_ids, pred_ids, labels=ordered_ids)
    precision, recall, f1, support = precision_recall_fscore_support(
        manual_ids,
        pred_ids,
        labels=ordered_ids,
        zero_division=0,
    )
    per_class_rows: List[Dict[str, Any]] = []
    for idx, label_id in enumerate(ordered_ids):
        per_class_rows.append(
            {
                "label_id": label_id,
                "class": DOWNSTREAM_LABELS[label_id],
                "precision": float(precision[idx]),
                "recall": float(recall[idx]),
                "f1": float(f1[idx]),
                "support": int(support[idx]),
            }
        )

    summary = {
        "accuracy": float(accuracy_score(manual_ids, pred_ids)),
        "balanced_accuracy": float(balanced_accuracy_score(manual_ids, pred_ids)),
        "macro_f1": float(f1_score(manual_ids, pred_ids, labels=ordered_ids, average="macro", zero_division=0)),
        "labels": labels,
    }
    return summary, per_class_rows, cm


def save_downstream_confusion_matrix_plot(path: Path, labels: Sequence[str], cm: np.ndarray) -> bool:
    try:
        import matplotlib.pyplot as plt
    except Exception:
        print("[DOWNSTREAM][WARN] matplotlib unavailable; skipping confusion matrix PNG.")
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    cm = np.asarray(cm, dtype=np.int64)
    row_sums = cm.sum(axis=1, keepdims=True)
    cm_norm = np.divide(
        cm.astype(np.float64),
        np.where(row_sums == 0, 1, row_sums),
        out=np.zeros_like(cm, dtype=np.float64),
    )

    fig, ax = plt.subplots(figsize=(7.5, 6.0))
    ax.imshow(cm_norm, interpolation="nearest", cmap="Blues", vmin=0.0, vmax=1.0)

    ax.set(
        xticks=np.arange(len(labels)),
        yticks=np.arange(len(labels)),
        xticklabels=list(labels),
        yticklabels=list(labels),
        xlabel="Predicted quality Assessment",
        ylabel="Manual quality assessment",
        title="Downstream Quality Assessment confusion matrix",
    )
    plt.setp(ax.get_xticklabels(), rotation=25, ha="right", rotation_mode="anchor")

    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            count = int(cm[i, j])
            frac = float(cm_norm[i, j])
            text = f"{count}\n{frac * 100.0:.1f}%" if row_sums[i, 0] > 0 else str(count)
            ax.text(
                j,
                i,
                text,
                ha="center",
                va="center",
                color="white" if frac > 0.5 else "black",
                fontsize=10,
            )

    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return True


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Downstream full-FOV quality evaluation.")
    parser.add_argument("--preset", default=ACTIVE_PRESET, choices=sorted(DOWNSTREAM_PRESETS))
    parser.add_argument("--checkpoint", default="")
    parser.add_argument("--images-root", default="")
    parser.add_argument("--manual-labels", default=DEFAULT_MANUAL_LABELS)
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--target-name", default="")
    parser.add_argument("--model-class", default="auto", choices=["auto", "RFDETRSmall", "RFDETRMedium", "RFDETRLarge"])
    parser.add_argument("--model-resolution", type=int, default=None)
    parser.add_argument("--score-floor", type=float, default=0.001)
    parser.add_argument("--score-threshold", type=float, default=0.30)
    parser.add_argument("--class-score-thresholds", default=fov_eval.default_class_score_threshold_string())
    parser.add_argument("--overlay-selection", default=OVERLAY_SELECTION)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--progress-every", type=int, default=10)
    parser.add_argument("--slice-height", type=int, default=672)
    parser.add_argument("--slice-width", type=int, default=672)
    parser.add_argument("--overlap-height-ratio", type=float, default=0.20)
    parser.add_argument("--overlap-width-ratio", type=float, default=0.20)
    parser.add_argument("--perform-standard-pred", action="store_true", default=True)
    parser.add_argument("--no-standard-pred", dest="perform_standard_pred", action="store_false")
    parser.add_argument("--postprocess-type", default="GREEDYNMM", choices=["GREEDYNMM", "NMM", "NMS", "LSNMS"])
    parser.add_argument("--postprocess-match-metric", default="IOU", choices=["IOU", "IOS"])
    parser.add_argument("--postprocess-match-threshold", type=float, default=0.50)
    parser.add_argument("--postprocess-class-agnostic", action="store_true")
    parser.add_argument("--cross-class-duplicate-suppression", action="store_true", default=CROSS_CLASS_DUPLICATE_SUPPRESSION)
    parser.add_argument("--disable-cross-class-duplicate-suppression", dest="cross_class_duplicate_suppression", action="store_false")
    parser.add_argument("--cross-class-duplicate-ios-threshold", type=float, default=CROSS_CLASS_DUPLICATE_IOS_THRESHOLD)
    parser.add_argument("--cross-class-duplicate-iou-threshold", type=float, default=CROSS_CLASS_DUPLICATE_IOU_THRESHOLD)
    parser.add_argument("--cross-class-duplicate-area-ratio-threshold", type=float, default=CROSS_CLASS_DUPLICATE_AREA_RATIO_THRESHOLD)
    return parser


def resolve_config(args: argparse.Namespace) -> DownstreamConfig:
    preset = _resolve_preset(args.preset)
    checkpoint = _optional_path(args.checkpoint) or Path(preset["checkpoint"])
    images_root = _optional_path(args.images_root) or Path(preset["images_root"])
    manual_labels_path = _optional_path(args.manual_labels)
    if manual_labels_path is None:
        raise ValueError("--manual-labels is required.")
    base_output_dir = Path(args.output_dir).expanduser()
    target_name = (args.target_name or "").strip() or str(preset.get("target_name", "Downstream"))
    output_dir, output_root_dir, run_timestamp = build_output_dir(base_output_dir, target_name)
    class_names = list(preset.get("class_names", fov_eval.DEFAULT_CLASS_NAMES))
    class_score_thresholds = fov_eval.parse_class_score_thresholds(
        str(args.class_score_thresholds),
        class_names,
        float(args.score_threshold),
    )

    model_class = args.model_class
    if model_class == "auto":
        model_class = local_eval.infer_model_class(checkpoint.parent, checkpoint)
    model_resolution = args.model_resolution if args.model_resolution is not None else local_eval.infer_model_resolution(checkpoint)

    return DownstreamConfig(
        preset_name=args.preset,
        checkpoint=checkpoint,
        images_root=images_root,
        manual_labels_path=manual_labels_path,
        sheet_name=(args.sheet_name or "").strip() or None,
        output_dir=output_dir,
        output_root_dir=output_root_dir,
        run_timestamp=run_timestamp,
        target_name=target_name,
        class_names=class_names,
        class_score_thresholds=class_score_thresholds,
        model_class=model_class,
        model_resolution=model_resolution,
        score_floor=float(args.score_floor),
        score_threshold=float(args.score_threshold),
        overlay_selection=str(args.overlay_selection),
        num_overlays=fov_eval.parse_overlay_selection(args.overlay_selection),
        seed=int(args.seed),
        progress_every=max(1, int(args.progress_every)),
        slice_height=int(args.slice_height) if args.slice_height else None,
        slice_width=int(args.slice_width) if args.slice_width else None,
        overlap_height_ratio=float(args.overlap_height_ratio),
        overlap_width_ratio=float(args.overlap_width_ratio),
        perform_standard_pred=bool(args.perform_standard_pred),
        postprocess_type=str(args.postprocess_type),
        postprocess_match_metric=str(args.postprocess_match_metric),
        postprocess_match_threshold=float(args.postprocess_match_threshold),
        postprocess_class_agnostic=bool(args.postprocess_class_agnostic),
        cross_class_duplicate_suppression=bool(args.cross_class_duplicate_suppression),
        cross_class_duplicate_ios_threshold=float(args.cross_class_duplicate_ios_threshold),
        cross_class_duplicate_iou_threshold=float(args.cross_class_duplicate_iou_threshold),
        cross_class_duplicate_area_ratio_threshold=float(args.cross_class_duplicate_area_ratio_threshold),
    )


def run_downstream_eval(cfg: DownstreamConfig) -> None:
    ensure_deps()
    _, get_sliced_prediction, _ = fov_eval.import_sahi()

    rows = load_manual_rows(cfg.manual_labels_path, cfg.sheet_name)
    if not rows:
        raise RuntimeError(f"No rows found in manual label file: {cfg.manual_labels_path}")
    image_index = build_image_index(cfg.images_root)

    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    overlays_dir = cfg.output_dir / "overlays"
    random.seed(cfg.seed)

    print(f"[DOWNSTREAM] Preset={cfg.preset_name}")
    print(f"[DOWNSTREAM] Checkpoint={cfg.checkpoint}")
    print(f"[DOWNSTREAM] Images root={cfg.images_root}")
    print(f"[DOWNSTREAM] Manual labels={cfg.manual_labels_path}")
    print(f"[DOWNSTREAM] Rows to process={len(rows)}")
    print(f"[DOWNSTREAM] Output directory={cfg.output_dir}")
    print(f"[DOWNSTREAM] Cross-class duplicate suppression={cfg.cross_class_duplicate_suppression}")

    ckpt_num_classes, ckpt_class_names = fov_eval.infer_checkpoint_runtime_metadata(cfg.checkpoint)
    effective_class_names = list(ckpt_class_names or cfg.class_names)
    effective_num_classes = ckpt_num_classes if ckpt_num_classes is not None else len(effective_class_names)
    effective_class_score_thresholds = fov_eval.parse_class_score_thresholds(
        ";".join(f"{name}={cfg.class_score_thresholds.get(name, cfg.score_threshold)}" for name in effective_class_names),
        effective_class_names,
        cfg.score_threshold,
    )

    model = fov_eval.load_model_for_fov(
        cfg.model_class,
        cfg.checkpoint,
        cfg.model_resolution,
        effective_num_classes,
        effective_class_names,
    )
    sahi_model = fov_eval.build_sahi_model(model, effective_class_names, cfg.score_floor)

    indices = list(range(len(rows)))
    if cfg.num_overlays == 0:
        overlay_indices: set[int] = set()
    elif cfg.num_overlays < 0 or cfg.num_overlays >= len(rows):
        overlay_indices = set(indices)
    else:
        overlay_indices = set(random.sample(indices, cfg.num_overlays))

    result_rows: List[Dict[str, Any]] = []
    manual_ids: List[int] = []
    pred_ids: List[int] = []
    total_cross_class_duplicates_suppressed = 0

    for idx, row in enumerate(rows, start=1):
        if idx == 1 or (idx % cfg.progress_every) == 0 or idx == len(rows):
            pct = 100.0 * idx / max(1, len(rows))
            print(f"[DOWNSTREAM] Progress: {idx}/{len(rows)} ({pct:.1f}%)")

        image_path = resolve_image_path(str(row["source_value"]), cfg.images_root, image_index)
        prediction_result = get_sliced_prediction(
            str(image_path),
            detection_model=sahi_model,
            slice_height=cfg.slice_height,
            slice_width=cfg.slice_width,
            overlap_height_ratio=cfg.overlap_height_ratio,
            overlap_width_ratio=cfg.overlap_width_ratio,
            perform_standard_pred=cfg.perform_standard_pred,
            postprocess_type=cfg.postprocess_type,
            postprocess_match_metric=cfg.postprocess_match_metric,
            postprocess_match_threshold=cfg.postprocess_match_threshold,
            postprocess_class_agnostic=cfg.postprocess_class_agnostic,
            verbose=0,
        )

        pred_boxes_list: List[List[float]] = []
        pred_scores_list: List[float] = []
        pred_cls_list: List[int] = []
        for obj in prediction_result.object_prediction_list:
            x1, y1, x2, y2 = obj.bbox.to_xyxy()
            pred_boxes_list.append([float(x1), float(y1), float(x2), float(y2)])
            pred_scores_list.append(float(obj.score.value))
            pred_cls_list.append(int(obj.category.id))

        pred_boxes = np.asarray(pred_boxes_list, dtype=np.float32).reshape((-1, 4)) if pred_boxes_list else np.zeros((0, 4), dtype=np.float32)
        pred_scores = np.asarray(pred_scores_list, dtype=np.float32) if pred_scores_list else np.zeros((0,), dtype=np.float32)
        pred_cls = np.asarray(pred_cls_list, dtype=np.int64) if pred_cls_list else np.zeros((0,), dtype=np.int64)
        pred_cls = fov_eval.normalize_model_class_ids(pred_cls, effective_class_names)

        keep = fov_eval.per_class_keep_mask(
            pred_scores,
            pred_cls,
            effective_class_names,
            effective_class_score_thresholds,
            cfg.score_threshold,
        )
        kept_boxes = pred_boxes[keep]
        kept_scores = pred_scores[keep]
        kept_cls = pred_cls[keep]
        n_kept_before_duplicate_suppression = int(len(kept_boxes))
        n_cross_class_duplicates_suppressed = 0
        if cfg.cross_class_duplicate_suppression:
            kept_boxes, kept_scores, kept_cls, n_cross_class_duplicates_suppressed = suppress_cross_class_duplicates(
                kept_boxes,
                kept_scores,
                kept_cls,
                effective_class_names,
                cfg.cross_class_duplicate_ios_threshold,
                cfg.cross_class_duplicate_iou_threshold,
                cfg.cross_class_duplicate_area_ratio_threshold,
            )
            total_cross_class_duplicates_suppressed += int(n_cross_class_duplicates_suppressed)

        count_map = count_class_predictions(kept_cls, effective_class_names)
        leu_count = int(count_map.get("Leucocyte", 0))
        epi_count = int(count_map.get("Squamous Epithelial Cell", 0))
        predicted_label_id, leu_score, epi_score, total_score = classify_quality_from_counts(leu_count, epi_count)
        manual_label_id = int(row["manual_label_id"])

        rel_path = fov_eval.safe_relpath(image_path, cfg.images_root)
        result_rows.append(
            {
                "file_name": rel_path,
                "image_name": image_path.name,
                "manual_label_id": manual_label_id,
                "manual_label": DOWNSTREAM_LABELS[manual_label_id],
                "predicted_label_id": predicted_label_id,
                "predicted_label": DOWNSTREAM_LABELS[predicted_label_id],
                "n_leucocyte": leu_count,
                "n_squamous_epithelial_cell": epi_count,
                "leucocyte_score": leu_score,
                "squamous_epithelial_score": epi_score,
                "total_quality_score": total_score,
                "n_predictions_raw": int(len(pred_boxes)),
                "n_predictions_kept_before_duplicate_suppression": n_kept_before_duplicate_suppression,
                "n_cross_class_duplicates_suppressed": int(n_cross_class_duplicates_suppressed),
                "n_predictions_kept": int(len(kept_boxes)),
            }
        )
        manual_ids.append(manual_label_id)
        pred_ids.append(predicted_label_id)

        if (idx - 1) in overlay_indices:
            draw_downstream_overlay(
                image_path=image_path,
                pred_boxes=kept_boxes,
                pred_scores=kept_scores,
                pred_cls=kept_cls,
                class_names=effective_class_names,
                leu_count=leu_count,
                epi_count=epi_count,
                manual_label_name=DOWNSTREAM_LABELS[manual_label_id],
                predicted_label_name=DOWNSTREAM_LABELS[predicted_label_id],
                output_path=overlays_dir / f"{image_path.stem}.png",
            )

    summary, per_class_rows, cm = compute_downstream_metrics(manual_ids, pred_ids)
    fieldnames = list(result_rows[0].keys()) if result_rows else []
    write_csv(cfg.output_dir / "downstream_per_image.csv", fieldnames, result_rows)
    write_csv(
        cfg.output_dir / "downstream_per_class_metrics.csv",
        ["label_id", "class", "precision", "recall", "f1", "support"],
        per_class_rows,
    )

    cm_labels = [DOWNSTREAM_LABELS[i] for i in [1, 2, 3]]
    local_eval.save_confusion_csv(cfg.output_dir / "downstream_confusion_matrix.csv", cm_labels, cm)
    save_downstream_confusion_matrix_plot(cfg.output_dir / "downstream_confusion_matrix.png", cm_labels, cm)

    json_dump(
        cfg.output_dir / "downstream_summary.json",
        {
            "target_name": cfg.target_name,
            "manual_labels_path": str(cfg.manual_labels_path),
            "images_root": str(cfg.images_root),
            "checkpoint": str(cfg.checkpoint),
            "class_names": effective_class_names,
            "class_score_thresholds": effective_class_score_thresholds,
            "quality_rule": {
                "leucocyte_score": {
                    "<10": -1,
                    "10-25": 0,
                    "26-50": 1,
                    ">50": 2,
                },
                "squamous_epithelial_score": {
                    "<10": 0,
                    "10-25": -1,
                    ">25": -2,
                },
                "classification": {
                    "Qualified": "total_score >= 1 and squamous_epithelial_count < 10",
                    "Partially qualified": "all other intermediate combinations",
                    "Not qualified": "total_score <= -1",
                },
            },
            "postprocess": {
                "type": cfg.postprocess_type,
                "match_metric": cfg.postprocess_match_metric,
                "match_threshold": cfg.postprocess_match_threshold,
                "class_agnostic": cfg.postprocess_class_agnostic,
            },
            "cross_class_duplicate_suppression": {
                "enabled": cfg.cross_class_duplicate_suppression,
                "ios_threshold": cfg.cross_class_duplicate_ios_threshold,
                "iou_threshold": cfg.cross_class_duplicate_iou_threshold,
                "area_ratio_threshold": cfg.cross_class_duplicate_area_ratio_threshold,
                "total_suppressed": total_cross_class_duplicates_suppressed,
            },
            "n_images": len(result_rows),
            "metrics": summary,
        },
    )


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    cfg = resolve_config(args)
    run_downstream_eval(cfg)


if __name__ == "__main__":
    main()
