# -*- coding: utf-8 -*-
"""
Create/update a Full FOV image-quality Excel overview from CVAT exports.

Purpose:
- Read the latest CVAT Full FOV export under BASE_DIR.
- Extract each image name and its image-level tag(s).
- Write a timestamped Excel workbook with a compact Master sheet:
  Image, Tag, Annotator, Comments.
- On later runs, copy the newest prior Full FOV workbook first so existing
  comments, cell fills/color markings, and other manual Excel edits are kept.
- Append newly exported images without overwriting manual review columns.

Supported CVAT exports:
- CVAT XML for images, e.g. annotations.xml
- CVAT JSON-like exports with an "images" list and image-level "tags"

Default output:
  BASE_DIR / Full_FOV_Master_Review_YYYY-MM-DD_HHMM.xlsx

Usage:
  python Helpers_General/FOVAnnotationOverviewGenerator_Master.py
  python Helpers_General/FOVAnnotationOverviewGenerator_Master.py --dry-run
  python Helpers_General/FOVAnnotationOverviewGenerator_Master.py --input "D:\\...\\annotations.xml"
  python Helpers_General/FOVAnnotationOverviewGenerator_Master.py --drop-missing
"""

from __future__ import annotations

import argparse
import copy
import json
import re
import shutil
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ================= CONFIG =================

BASE_DIR = Path(
    r"D:\PHD\PhdData\CellScanData\Annotation_Backups\Quality Assessment Full FOV Backups"
)

MASTER_PREFIX = "Full_FOV_Master_Review"
MASTER_SHEET = "Master"
GENERATED_SHEETS = ["Summary", "Tag_Totals", "Labels"]

SOURCE_COLUMNS = ["Image", "Tag"]
MANUAL_COLUMNS = ["Annotator", "Comments"]
MASTER_COLUMNS = ["Image", "Tag", "Annotator", "Comments"]

EXPORT_EXTENSIONS = {".json", ".xml"}


# ==========================================


@dataclass(frozen=True)
class CvatRecord:
    image: str
    tag: str
    sample: str = ""
    task: str = ""
    task_id: str = ""
    image_id: str = ""
    width: str = ""
    height: str = ""
    subset: str = ""
    tag_source: str = ""


@dataclass(frozen=True)
class ParsedExport:
    path: Path
    format_name: str
    records: List[CvatRecord]
    labels: Dict[str, str]


def timestamp_name() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H%M")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null"}:
        return ""
    return text


def normalize_key(value: Any) -> str:
    return normalize_text(value).casefold()


def first_nonempty(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def natural_key(text: str) -> List[Any]:
    parts = re.split(r"(\d+)", normalize_text(text))
    return [int(part) if part.isdigit() else part.casefold() for part in parts]


def extract_sample_name(*values: str) -> str:
    for value in values:
        match = re.search(r"\bSample\s*(\d+)\b", normalize_text(value), flags=re.IGNORECASE)
        if match:
            return f"Sample{int(match.group(1))}"
    return ""


def unique_join(values: Iterable[Any]) -> str:
    seen = set()
    out: List[str] = []
    for value in values:
        text = normalize_text(value)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return "; ".join(out)


def color_to_argb(color: str) -> str:
    color = normalize_text(color).lstrip("#")
    if len(color) == 6 and re.fullmatch(r"[0-9A-Fa-f]{6}", color):
        return f"FF{color.upper()}"
    if len(color) == 8 and re.fullmatch(r"[0-9A-Fa-f]{8}", color):
        return color.upper()
    return ""


def find_files(root: Path, extensions: Sequence[str]) -> List[Path]:
    if root.is_file():
        return [root] if root.suffix.lower() in extensions else []
    files = [
        path
        for path in root.rglob("*")
        if path.is_file() and path.suffix.lower() in extensions and not path.name.startswith("~$")
    ]
    files.sort(key=lambda path: path.stat().st_mtime)
    return files


def find_latest_export(base_dir: Path) -> Path:
    files = find_files(base_dir, tuple(EXPORT_EXTENSIONS))
    if not files:
        raise FileNotFoundError(f"No CVAT .json/.xml export found under {base_dir}")
    return files[-1]


def find_prior_masters(base_dir: Path) -> List[Path]:
    files = [
        path
        for path in base_dir.rglob(f"{MASTER_PREFIX}*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    files.sort(key=lambda path: path.stat().st_mtime)
    return files


def parse_xml_export(path: Path) -> ParsedExport:
    root = ET.parse(path).getroot()

    labels: Dict[str, str] = {}
    for label_el in root.findall(".//meta//label"):
        name = normalize_text(label_el.findtext("name"))
        color = normalize_text(label_el.findtext("color"))
        label_type = normalize_text(label_el.findtext("type"))
        if name and (not label_type or label_type.casefold() == "tag"):
            labels[name] = color

    task_id_to_name: Dict[str, str] = {}
    for task_el in root.findall(".//meta//task"):
        task_id = normalize_text(task_el.findtext("id"))
        task_name = normalize_text(task_el.findtext("name"))
        if task_id:
            task_id_to_name[task_id] = task_name

    records: List[CvatRecord] = []
    for image_el in root.findall(".//image"):
        image_name = normalize_text(image_el.attrib.get("name"))
        if not image_name:
            continue

        task_id = normalize_text(image_el.attrib.get("task_id"))
        task_name = task_id_to_name.get(task_id, "")
        tag_labels: List[str] = []
        tag_sources: List[str] = []
        for tag_el in image_el.findall("tag"):
            tag_labels.append(tag_el.attrib.get("label", ""))
            tag_sources.append(tag_el.attrib.get("source", ""))

        records.append(
            CvatRecord(
                image=image_name,
                tag=unique_join(tag_labels),
                sample=extract_sample_name(task_name, image_name),
                task=task_name,
                task_id=task_id,
                image_id=normalize_text(image_el.attrib.get("id")),
                width=normalize_text(image_el.attrib.get("width")),
                height=normalize_text(image_el.attrib.get("height")),
                subset=normalize_text(image_el.attrib.get("subset")),
                tag_source=unique_join(tag_sources),
            )
        )

    return ParsedExport(path=path, format_name="xml", records=sort_records(records), labels=labels)


def get_nested_dict(data: Dict[str, Any], keys: Sequence[str]) -> Dict[str, Any]:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict):
            return {}
        current = current.get(key, {})
    return current if isinstance(current, dict) else {}


def collect_json_labels(data: Any) -> Dict[str, str]:
    labels: Dict[str, str] = {}

    def add_label(label: Any) -> None:
        if not isinstance(label, dict):
            return
        name = first_nonempty(label.get("name"), label.get("label"))
        if not name:
            return
        labels[name] = normalize_text(label.get("color"))

    if isinstance(data, dict):
        for label in data.get("labels", []) or []:
            add_label(label)
        for category in data.get("categories", []) or []:
            add_label(category)
        for label in get_nested_dict(data, ["meta", "project"]).get("labels", []) or []:
            add_label(label)
        for label in get_nested_dict(data, ["meta", "task"]).get("labels", []) or []:
            add_label(label)
    return labels


def collect_json_label_id_maps(data: Any) -> Tuple[Dict[str, str], Dict[str, str]]:
    label_id_to_name: Dict[str, str] = {}
    category_id_to_name: Dict[str, str] = {}

    def add_label(label: Any, target: Dict[str, str]) -> None:
        if not isinstance(label, dict):
            return
        label_id = first_nonempty(label.get("id"), label.get("label_id"), label.get("category_id"))
        name = first_nonempty(label.get("name"), label.get("label"))
        if label_id and name:
            target[label_id] = name

    if isinstance(data, dict):
        for label in data.get("labels", []) or []:
            add_label(label, label_id_to_name)
        for label in get_nested_dict(data, ["meta", "project"]).get("labels", []) or []:
            add_label(label, label_id_to_name)
        for label in get_nested_dict(data, ["meta", "task"]).get("labels", []) or []:
            add_label(label, label_id_to_name)
        for category in data.get("categories", []) or []:
            add_label(category, category_id_to_name)

    return label_id_to_name, category_id_to_name


def collect_json_task_map(data: Any) -> Dict[str, str]:
    task_id_to_name: Dict[str, str] = {}
    if not isinstance(data, dict):
        return task_id_to_name
    tasks = get_nested_dict(data, ["meta", "project"]).get("tasks", []) or []
    if isinstance(tasks, list):
        for task in tasks:
            if not isinstance(task, dict):
                continue
            task_id = first_nonempty(task.get("id"), task.get("task_id"))
            task_name = first_nonempty(task.get("name"), task.get("task_name"))
            if task_id:
                task_id_to_name[task_id] = task_name
    return task_id_to_name


def parse_json_tag(
    tag: Any,
    label_id_to_name: Dict[str, str],
    category_id_to_name: Dict[str, str],
) -> Tuple[str, str]:
    if isinstance(tag, str):
        return normalize_text(tag), ""
    if not isinstance(tag, dict):
        return "", ""

    label = first_nonempty(
        tag.get("label"),
        tag.get("name"),
        tag.get("label_name"),
        tag.get("category_name"),
    )
    if not label:
        label_id = first_nonempty(tag.get("label_id"), tag.get("labelId"))
        category_id = first_nonempty(tag.get("category_id"), tag.get("categoryId"))
        label = label_id_to_name.get(label_id, "") or category_id_to_name.get(category_id, "")

    source = first_nonempty(tag.get("source"), tag.get("origin"))
    return label, source


def group_top_level_json_tags(
    data: Any,
    label_id_to_name: Dict[str, str],
    category_id_to_name: Dict[str, str],
) -> Dict[str, List[Tuple[str, str]]]:
    grouped: Dict[str, List[Tuple[str, str]]] = {}
    if not isinstance(data, dict):
        return grouped

    for tag in data.get("tags", []) or []:
        if not isinstance(tag, dict):
            continue
        image_id = first_nonempty(tag.get("image_id"), tag.get("imageId"), tag.get("image"))
        if not image_id:
            continue
        grouped.setdefault(image_id, []).append(parse_json_tag(tag, label_id_to_name, category_id_to_name))

    # Some exports can place image-level tags in annotations. Avoid treating
    # COCO bounding boxes as tags by requiring an explicit tag marker.
    for ann in data.get("annotations", []) or []:
        if not isinstance(ann, dict):
            continue
        ann_type = first_nonempty(ann.get("type"), ann.get("shape_type"), ann.get("kind"))
        if ann_type.casefold() != "tag":
            continue
        image_id = first_nonempty(ann.get("image_id"), ann.get("imageId"), ann.get("image"))
        if not image_id:
            continue
        grouped.setdefault(image_id, []).append(parse_json_tag(ann, label_id_to_name, category_id_to_name))

    return grouped


def get_json_images(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if not isinstance(data, dict):
        return []
    images = data.get("images", [])
    if isinstance(images, list):
        return [item for item in images if isinstance(item, dict)]
    return []


def parse_json_export(path: Path) -> ParsedExport:
    with path.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)

    labels = collect_json_labels(data)
    label_id_to_name, category_id_to_name = collect_json_label_id_maps(data)
    task_id_to_name = collect_json_task_map(data)
    top_level_tags = group_top_level_json_tags(data, label_id_to_name, category_id_to_name)

    records: List[CvatRecord] = []
    for image in get_json_images(data):
        image_name = first_nonempty(image.get("name"), image.get("file_name"), image.get("filename"))
        if not image_name:
            continue

        image_id = first_nonempty(image.get("id"), image.get("image_id"), image.get("imageId"))
        task_id = first_nonempty(image.get("task_id"), image.get("taskId"))
        task_name = first_nonempty(image.get("task_name"), image.get("task"), task_id_to_name.get(task_id))

        raw_tags: List[Any] = []
        for key in ("tags", "labels"):
            value = image.get(key)
            if isinstance(value, list):
                raw_tags.extend(value)
        parsed_tags = [
            parse_json_tag(tag, label_id_to_name, category_id_to_name)
            for tag in raw_tags
        ]
        parsed_tags.extend(top_level_tags.get(image_id, []))

        records.append(
            CvatRecord(
                image=image_name,
                tag=unique_join(label for label, _ in parsed_tags),
                sample=extract_sample_name(task_name, image_name),
                task=task_name,
                task_id=task_id,
                image_id=image_id,
                width=first_nonempty(image.get("width"), image.get("w")),
                height=first_nonempty(image.get("height"), image.get("h")),
                subset=first_nonempty(image.get("subset")),
                tag_source=unique_join(source for _, source in parsed_tags),
            )
        )

    return ParsedExport(path=path, format_name="json", records=sort_records(records), labels=labels)


def parse_export(path: Path) -> ParsedExport:
    suffix = path.suffix.lower()
    if suffix == ".xml":
        return parse_xml_export(path)
    if suffix == ".json":
        return parse_json_export(path)
    raise ValueError(f"Unsupported export format: {path}")


def sort_records(records: Sequence[CvatRecord]) -> List[CvatRecord]:
    return sorted(records, key=lambda record: (natural_key(record.sample), natural_key(record.image)))


def make_output_path(base_dir: Path) -> Path:
    output = base_dir / f"{MASTER_PREFIX}_{timestamp_name()}.xlsx"
    counter = 2
    while output.exists():
        output = base_dir / f"{MASTER_PREFIX}_{timestamp_name()}_{counter:02d}.xlsx"
        counter += 1
    return output


def prepare_workbook(output_path: Path, prior_master: Optional[Path]) -> Tuple[Any, Worksheet]:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if prior_master is not None:
        shutil.copy2(prior_master, output_path)
        wb = load_workbook(output_path)
        if MASTER_SHEET in wb.sheetnames:
            ws = wb[MASTER_SHEET]
        else:
            ws = wb.active
            ws.title = MASTER_SHEET
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_SHEET
    for col_idx, header in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell)
    return wb, ws


def apply_header_style(cell: Any) -> None:
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="FF1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_default_cell_style(cell: Any) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def header_map(ws: Worksheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for cell in ws[1]:
        name = normalize_text(cell.value)
        if name and name not in headers:
            headers[name] = int(cell.column)
    return headers


def ensure_master_headers(ws: Worksheet) -> Dict[str, int]:
    if ws.max_row == 1 and ws.max_column == 1 and not normalize_text(ws.cell(1, 1).value):
        for col_idx, header in enumerate(MASTER_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            apply_header_style(cell)

    headers = header_map(ws)
    for header in MASTER_COLUMNS:
        if header in headers:
            continue
        col_idx = ws.max_column + 1
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell)
        headers[header] = col_idx
    return headers


def prune_master_columns(ws: Worksheet, keep_columns: Sequence[str]) -> Dict[str, int]:
    """Drop obsolete columns after headers exist, preserving kept cell styles/comments."""
    headers = header_map(ws)
    keep = set(keep_columns)
    delete_indices = [
        col_idx
        for header, col_idx in headers.items()
        if header not in keep
    ]
    for col_idx in sorted(delete_indices, reverse=True):
        ws.delete_cols(col_idx, 1)
    return header_map(ws)


def existing_rows_by_image(ws: Worksheet, headers: Dict[str, int]) -> Tuple[Dict[str, int], List[str]]:
    image_col = headers["Image"]
    rows: Dict[str, int] = {}
    duplicates: List[str] = []
    for row_idx in range(2, ws.max_row + 1):
        image = normalize_text(ws.cell(row=row_idx, column=image_col).value)
        if not image:
            continue
        key = normalize_key(image)
        if key in rows:
            duplicates.append(image)
            continue
        rows[key] = row_idx
    return rows, duplicates


def record_to_row_values(record: CvatRecord, export_path: Path, export_modified: str) -> Dict[str, str]:
    return {
        "Image": record.image,
        "Tag": record.tag,
        "Present In Latest": "Yes",
        "Sample": record.sample,
        "Task": record.task,
        "Task ID": record.task_id,
        "CVAT Image ID": record.image_id,
        "Width": record.width,
        "Height": record.height,
        "Subset": record.subset,
        "Tag Source": record.tag_source,
        "Export File": str(export_path),
        "Export Modified": export_modified,
        "Reviewed": "",
        "Annotator": "",
        "Comments": "",
    }


def write_source_values(
    ws: Worksheet,
    row_idx: int,
    headers: Dict[str, int],
    values: Dict[str, str],
) -> None:
    for header in SOURCE_COLUMNS:
        ws.cell(row=row_idx, column=headers[header]).value = values.get(header, "")


def append_record_row(
    ws: Worksheet,
    headers: Dict[str, int],
    values: Dict[str, str],
    labels: Dict[str, str],
    apply_tag_colors: bool,
) -> int:
    row_idx = ws.max_row + 1
    for header in MASTER_COLUMNS:
        cell = ws.cell(row=row_idx, column=headers[header], value=values.get(header, ""))
        apply_default_cell_style(cell)

    if apply_tag_colors:
        apply_tag_fill(ws.cell(row=row_idx, column=headers["Tag"]), values.get("Tag", ""), labels)
    return row_idx


def apply_tag_fill(cell: Any, tag_value: str, labels: Dict[str, str]) -> None:
    first_tag = normalize_text(tag_value).split(";")[0].strip()
    if not first_tag:
        return
    color = labels.get(first_tag, "")
    argb = color_to_argb(color)
    if not argb:
        return
    cell.fill = PatternFill(fill_type="solid", fgColor=argb)


def mark_missing_rows(
    ws: Worksheet,
    headers: Dict[str, int],
    latest_keys: set[str],
    drop_missing: bool,
) -> int:
    existing_rows, _ = existing_rows_by_image(ws, headers)
    missing_row_indices = [
        row_idx
        for key, row_idx in existing_rows.items()
        if key not in latest_keys
    ]

    if drop_missing:
        for row_idx in sorted(missing_row_indices, reverse=True):
            ws.delete_rows(row_idx, 1)
        return len(missing_row_indices)

    return len(missing_row_indices)


def format_master_sheet(ws: Worksheet, headers: Dict[str, int]) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    widths = {
        "Image": 38,
        "Tag": 24,
        "Annotator": 14,
        "Comments": 45,
    }
    for header, width in widths.items():
        if header in headers:
            ws.column_dimensions[get_column_letter(headers[header])].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = copy.copy(cell.alignment)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )


def replace_generated_sheet(wb: Any, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_summary_sheets(
    wb: Any,
    parsed: ParsedExport,
    output_path: Path,
    prior_master: Optional[Path],
    updated_count: int,
    appended_count: int,
    missing_count: int,
    drop_missing: bool,
) -> None:
    summary = replace_generated_sheet(wb, "Summary")
    summary_rows = [
        ("Metric", "Value"),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Output Workbook", str(output_path)),
        ("Prior Workbook", str(prior_master) if prior_master is not None else ""),
        ("CVAT Export", str(parsed.path)),
        ("CVAT Export Modified", datetime.fromtimestamp(parsed.path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")),
        ("CVAT Export Format", parsed.format_name),
        ("Images In Latest Export", len(parsed.records)),
        ("Existing Rows Updated", updated_count),
        ("New Rows Appended", appended_count),
        ("Missing Old Rows", missing_count),
        ("Missing Old Rows Handling", "Dropped" if drop_missing else "Retained"),
    ]
    for row_idx, row in enumerate(summary_rows, start=1):
        summary.cell(row=row_idx, column=1, value=row[0])
        summary.cell(row=row_idx, column=2, value=row[1])
    apply_header_style(summary.cell(row=1, column=1))
    apply_header_style(summary.cell(row=1, column=2))
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 90

    tag_totals = replace_generated_sheet(wb, "Tag_Totals")
    tag_totals.cell(row=1, column=1, value="Tag")
    tag_totals.cell(row=1, column=2, value="Image Count")
    apply_header_style(tag_totals.cell(row=1, column=1))
    apply_header_style(tag_totals.cell(row=1, column=2))
    counts = Counter(record.tag or "(No tag)" for record in parsed.records)
    for row_idx, (tag, count) in enumerate(sorted(counts.items(), key=lambda item: natural_key(item[0])), start=2):
        tag_totals.cell(row=row_idx, column=1, value=tag)
        tag_totals.cell(row=row_idx, column=2, value=count)
    tag_totals.column_dimensions["A"].width = 28
    tag_totals.column_dimensions["B"].width = 14

    labels = replace_generated_sheet(wb, "Labels")
    labels.cell(row=1, column=1, value="Label")
    labels.cell(row=1, column=2, value="CVAT Color")
    apply_header_style(labels.cell(row=1, column=1))
    apply_header_style(labels.cell(row=1, column=2))
    for row_idx, (label, color) in enumerate(sorted(parsed.labels.items(), key=lambda item: natural_key(item[0])), start=2):
        labels.cell(row=row_idx, column=1, value=label)
        labels.cell(row=row_idx, column=2, value=color)
        argb = color_to_argb(color)
        if argb:
            labels.cell(row=row_idx, column=1).fill = PatternFill(fill_type="solid", fgColor=argb)
            labels.cell(row=row_idx, column=2).fill = PatternFill(fill_type="solid", fgColor=argb)
    labels.column_dimensions["A"].width = 28
    labels.column_dimensions["B"].width = 14


def update_workbook(
    parsed: ParsedExport,
    output_path: Path,
    prior_master: Optional[Path],
    drop_missing: bool,
    apply_tag_colors: bool,
) -> Dict[str, Any]:
    wb, ws = prepare_workbook(output_path, prior_master)
    headers = ensure_master_headers(ws)
    existing_rows, duplicates = existing_rows_by_image(ws, headers)

    export_modified = datetime.fromtimestamp(parsed.path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    updated_count = 0
    appended_count = 0
    latest_keys: set[str] = set()

    for record in parsed.records:
        key = normalize_key(record.image)
        latest_keys.add(key)
        values = record_to_row_values(record, parsed.path, export_modified)
        if key in existing_rows:
            write_source_values(ws, existing_rows[key], headers, values)
            updated_count += 1
        else:
            append_record_row(ws, headers, values, parsed.labels, apply_tag_colors)
            appended_count += 1

    missing_count = mark_missing_rows(ws, headers, latest_keys, drop_missing)
    headers = prune_master_columns(ws, MASTER_COLUMNS)
    headers = ensure_master_headers(ws)
    format_master_sheet(ws, headers)

    write_summary_sheets(
        wb=wb,
        parsed=parsed,
        output_path=output_path,
        prior_master=prior_master,
        updated_count=updated_count,
        appended_count=appended_count,
        missing_count=missing_count,
        drop_missing=drop_missing,
    )

    wb.save(output_path)
    return {
        "updated": updated_count,
        "appended": appended_count,
        "missing": missing_count,
        "duplicates": duplicates,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Create/update a Full FOV CVAT image-tag Excel overview."
    )
    parser.add_argument("--base-dir", default=str(BASE_DIR), help="Root folder containing CVAT backups and master workbooks.")
    parser.add_argument("--input", default="", help="Specific CVAT .json/.xml export or folder to scan.")
    parser.add_argument("--output", default="", help="Specific output .xlsx path. Default creates a timestamped workbook.")
    parser.add_argument("--prior-master", default="", help="Specific prior workbook to copy/update. Default uses newest prior Full FOV master.")
    parser.add_argument("--drop-missing", action="store_true", help="Drop old workbook rows not present in the latest CVAT export.")
    parser.add_argument("--no-tag-colors", action="store_true", help="Do not apply CVAT label colors to newly added tag cells.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report planned changes without writing a workbook.")
    return parser


def resolve_path(raw: str) -> Optional[Path]:
    text = normalize_text(raw)
    if not text:
        return None
    return Path(text).expanduser()


def main() -> None:
    args = build_parser().parse_args()

    base_dir = Path(args.base_dir).expanduser()
    input_path = resolve_path(args.input)
    export_path = find_latest_export(input_path) if input_path else find_latest_export(base_dir)
    parsed = parse_export(export_path)
    if not parsed.records:
        raise RuntimeError(f"No image records found in CVAT export: {export_path}")

    prior_master = resolve_path(args.prior_master)
    prior_masters: List[Path] = []
    if prior_master is None:
        prior_masters = find_prior_masters(base_dir)
        prior_master = prior_masters[-1] if prior_masters else None
    elif not prior_master.exists():
        raise FileNotFoundError(f"Prior workbook not found: {prior_master}")

    output_path = resolve_path(args.output) or make_output_path(base_dir)
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Output must be an .xlsx file: {output_path}")

    existing_count = 0
    missing_count = 0
    appended_count = len(parsed.records)
    if prior_master is not None and prior_master.exists():
        wb = load_workbook(prior_master, read_only=True, data_only=False)
        sheet_name = MASTER_SHEET if MASTER_SHEET in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = header_map(ws) if ws.max_row else {}
        if "Image" in headers:
            existing_keys = {
                normalize_key(ws.cell(row=row_idx, column=headers["Image"]).value)
                for row_idx in range(2, ws.max_row + 1)
                if normalize_text(ws.cell(row=row_idx, column=headers["Image"]).value)
            }
            latest_keys = {normalize_key(record.image) for record in parsed.records}
            existing_count = len(existing_keys & latest_keys)
            appended_count = len(latest_keys - existing_keys)
            missing_count = len(existing_keys - latest_keys)
        wb.close()

    print("== Full FOV CVAT overview ==")
    print(f"CVAT export:       {parsed.path}")
    print(f"Export format:     {parsed.format_name}")
    print(f"Images in export:  {len(parsed.records)}")
    print(f"Prior workbook:    {prior_master if prior_master else '(none)'}")
    print(f"Output workbook:   {output_path}")
    print(f"Existing updated:  {existing_count}")
    print(f"New appended:      {appended_count}")
    print(f"Missing old rows:  {missing_count} ({'drop' if args.drop_missing else 'retain'})")

    if args.dry_run:
        print("\n[DRY RUN] No workbook written.")
        return

    stats = update_workbook(
        parsed=parsed,
        output_path=output_path,
        prior_master=prior_master,
        drop_missing=bool(args.drop_missing),
        apply_tag_colors=not bool(args.no_tag_colors),
    )

    if stats["duplicates"]:
        print("\nWarning: duplicate image rows were found in the prior workbook.")
        print("Only the first row for each duplicate image was updated.")
        for image in stats["duplicates"][:20]:
            print(f"- {image}")
        if len(stats["duplicates"]) > 20:
            print(f"- ... {len(stats['duplicates']) - 20} more")

    print(f"\n[OK] Workbook written: {output_path}")
    print(
        "Rows updated/appended/missing: "
        f"{stats['updated']}/{stats['appended']}/{stats['missing']}"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise
